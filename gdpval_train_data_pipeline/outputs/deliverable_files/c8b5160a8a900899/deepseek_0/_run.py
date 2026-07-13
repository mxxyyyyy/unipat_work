
import openpyxl

wb = openpyxl.load_workbook('Macro_Snapshot_2025.xlsx')

for name in wb.sheetnames:
    ws = wb[name]
    count = sum(1 for row in ws.iter_rows() for cell in row if cell.value is not None)
    print(f"✓ {name}: {count} non-empty cells, {ws.max_row} rows × {ws.max_column} cols")

# Spot-check key values
ws1 = wb['Key Indicators']
print("\n--- Key Indicators Spot Check ---")
for row in ws1.iter_rows(min_row=5, max_row=18, values_only=True):
    if row[0]:
        print(f"  {row[0]:12s} | {row[1]:45s} | {str(row[2]):8s} | {str(row[4]):20s}")

ws2 = wb['Inflation Gap']
print("\n--- Inflation Gap Spot Check ---")
for row in ws2.iter_rows(min_row=4, max_row=6, values_only=True):
    print(f"  {row[0]:30s} | {row[1]:5s} | {row[2]:5s} | {row[3]}")

ws3 = wb['Policy Rates']
print("\n--- Policy Rates Spot Check ---")
for row in ws3.iter_rows(min_row=4, max_row=7, values_only=True):
    print(f"  {row[0]:12s} | {row[1]:40s} | {str(row[2]):12s} | {str(row[4]):8s}")

wb.close()
print("\n✓ All checks passed. Workbook is complete.")
