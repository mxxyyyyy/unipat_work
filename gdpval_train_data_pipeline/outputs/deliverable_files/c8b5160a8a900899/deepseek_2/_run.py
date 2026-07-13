
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Styles ──────────────────────────────────────────
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
sub_header_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
title_font = Font(name='Calibri', size=14, bold=True, color='1F3864')
data_font = Font(name='Calibri', size=11)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical='top')
center_align = Alignment(horizontal='center', vertical='top')

def style_header_row(ws, row, ncols):
    for col in range(1, ncols+1):
        c = ws.cell(row=row, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        c.border = thin_border

def style_data_rows(ws, start_row, end_row, ncols):
    for r in range(start_row, end_row+1):
        for col in range(1, ncols+1):
            c = ws.cell(row=r, column=col)
            c.font = data_font
            c.alignment = wrap_align
            c.border = thin_border

# ═══════════════════════════════════════════════════
# Sheet 1 — Key Indicators
# ═══════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Key Indicators"

# Title row
ws1.merge_cells('A1:F1')
ws1['A1'] = "Macro Snapshot — Key Indicators (from Fed & ECB Reports, Early 2025)"
ws1['A1'].font = title_font
ws1['A1'].alignment = Alignment(horizontal='left', vertical='center')
ws1.row_dimensions[1].height = 28

# Headers in row 3
headers1 = ['Economy', 'Indicator', 'Latest Value', 'Unit', 'As-of Date', 'Source File']
for ci, h in enumerate(headers1, 1):
    ws1.cell(row=3, column=ci, value=h)
style_header_row(ws1, 3, len(headers1))

# Data
key_data = [
    # US data
    ['United States', 'Headline PCE Inflation (YoY)', 2.1, '% YoY', 'Apr 2025',
     'Fed Monetary Policy Report (June 2025)'],
    ['United States', 'Headline PCE Inflation (YoY, prior)', 2.6, '% YoY', 'Dec 2024',
     'Fed Monetary Policy Report (Feb 2025)'],
    ['United States', 'Core PCE Inflation (YoY, excl. food & energy)', 2.5, '% YoY', 'Apr 2025',
     'Fed Monetary Policy Report (June 2025)'],
    ['United States', 'Core PCE Inflation (YoY, prior)', 2.9, '% YoY', 'Dec 2024',
     'Fed Monetary Policy Report (June 2025) — references end-2024'],
    ['United States', 'Dallas Fed Trimmed Mean PCE (YoY)', 2.5, '% YoY', 'Apr 2025',
     'Fed Monetary Policy Report (June 2025)'],
    ['United States', 'PCE Food Prices (YoY)', 1.9, '% YoY', 'Apr 2025',
     'Fed Monetary Policy Report (June 2025)'],
    ['United States', 'PCE Energy Prices (YoY)', -6.0, '% YoY (approx.)', 'Apr 2025',
     'Fed Monetary Policy Report (June 2025) — "drop of almost 6 percent"'],
    ['United States', 'Real GDP Growth (annual)', 2.5, '%', '2024 (full year)',
     'Fed Monetary Policy Report (Feb 2025)'],
    ['United States', 'Unemployment Rate (U-3)', 4.2, '%', 'May 2025',
     'Fed Monetary Policy Report (June 2025)'],
    ['United States', 'Unemployment Rate (U-3, prior)', 4.1, '%', 'Dec 2024',
     'Fed Monetary Policy Report (Feb 2025)'],
    ['United States', 'Fed Funds Target Range (upper bound)', 4.50, '%', 'Jun 2025',
     'Fed Monetary Policy Report (June 2025)'],
    ['United States', 'Fed Funds Target Range (lower bound)', 4.25, '%', 'Jun 2025',
     'Fed Monetary Policy Report (June 2025)'],

    # Euro Area data
    ['Euro Area', 'Headline HICP Inflation (YoY)', 2.4, '% YoY', 'Dec 2024',
     'ECB Economic Bulletin (Issue 1/2025)'],
    ['Euro Area', 'Headline HICP Inflation (flash estimate, YoY)', 2.5, '% YoY', 'Jan 2025',
     'ECB Economic Bulletin (Issue 1/2025) — footnote 5, flash estimate'],
    ['Euro Area', 'Services Inflation (HICP, YoY)', 4.0, '% YoY', 'Dec 2024',
     'ECB Economic Bulletin (Issue 1/2025)'],
    ['Euro Area', 'Food Price Inflation (HICP, YoY)', 2.6, '% YoY', 'Dec 2024',
     'ECB Economic Bulletin (Issue 1/2025)'],
    ['Euro Area', 'Non-Energy Industrial Goods Inflation (HICP, YoY)', 0.5, '% YoY', 'Dec 2024',
     'ECB Economic Bulletin (Issue 1/2025)'],
    ['Euro Area', 'GDP Growth', None, '(not explicitly stated)', '—',
     'ECB Economic Bulletin (Issue 1/2025) — advance Q4 2024 estimate published after cut-off'],
    ['Euro Area', 'Unemployment Rate', None, '(not explicitly stated)', '—',
     'ECB Economic Bulletin (Issue 1/2025) — Chart 5 references Nov 2024 but no value in text'],
    ['Euro Area', 'ECB Deposit Facility Rate', 2.75, '%', '5 Feb 2025 (effective)',
     'ECB Economic Bulletin (Issue 1/2025)'],
    ['Euro Area', 'ECB Main Refinancing Operations Rate', 2.90, '%', '5 Feb 2025 (effective)',
     'ECB Economic Bulletin (Issue 1/2025)'],
    ['Euro Area', 'ECB Marginal Lending Facility Rate', 3.15, '%', '5 Feb 2025 (effective)',
     'ECB Economic Bulletin (Issue 1/2025)'],
]

for ri, row in enumerate(key_data, 4):
    for ci, val in enumerate(row, 1):
        ws1.cell(row=ri, column=ci, value=val if val is not None else 'N/A')

style_data_rows(ws1, 4, 3 + len(key_data), len(headers1))

# column widths
col_widths_1 = [18, 48, 16, 22, 18, 52]
for ci, w in enumerate(col_widths_1, 1):
    ws1.column_dimensions[get_column_letter(ci)].width = w

# ═══════════════════════════════════════════════════
# Sheet 2 — Inflation Gap
# ═══════════════════════════════════════════════════
ws2 = wb.create_sheet("Inflation Gap")

ws2.merge_cells('A1:E1')
ws2['A1'] = "Inflation Gap — Deviation from Central Bank Targets"
ws2['A1'].font = title_font
ws2.row_dimensions[1].height = 28

headers2 = ['Economy', 'Latest Headline Inflation (%)', 'Inflation Target (%)', 'Deviation (pp)', 'Notes']
for ci, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=ci, value=h)
style_header_row(ws2, 3, len(headers2))

# US: latest PCE headline = 2.1%, target = 2.0%
# EA: latest HICP headline = 2.4% (Dec) / 2.5% (Jan flash), target = 2.0%
gap_data = [
    ['United States', 2.1, 2.0, 2.1 - 2.0,
     'PCE inflation Apr 2025 (from June 2025 Fed report). Target is FOMC 2% PCE objective.'],
    ['Euro Area', 2.4, 2.0, 2.4 - 2.0,
     'HICP inflation Dec 2024 (from ECB Bulletin Issue 1/2025). Target is ECB 2% HICP.'],
    ['Euro Area (Jan flash)', 2.5, 2.0, 2.5 - 2.0,
     'HICP flash estimate for Jan 2025 = 2.5% (footnote in ECB Bulletin).'],
]

for ri, row in enumerate(gap_data, 4):
    for ci, val in enumerate(row, 1):
        ws2.cell(row=ri, column=ci, value=val)

style_data_rows(ws2, 4, 3 + len(gap_data), len(headers2))

# Format deviation with sign
for ri in range(4, 4 + len(gap_data)):
    ws2.cell(row=ri, column=4).number_format = '+0.0;-0.0'

col_widths_2 = [18, 26, 20, 16, 75]
for ci, w in enumerate(col_widths_2, 1):
    ws2.column_dimensions[get_column_letter(ci)].width = w

# ═══════════════════════════════════════════════════
# Sheet 3 — Policy Rates
# ═══════════════════════════════════════════════════
ws3 = wb.create_sheet("Policy Rates")

ws3.merge_cells('A1:G1')
ws3['A1'] = "Policy Rates — Current Levels and Recent Changes"
ws3['A1'].font = title_font
ws3.row_dimensions[1].height = 28

headers3 = ['Economy', 'Rate Type', 'Current Level', 'Previous Level', 'Change (bp)', 'As-of Date', 'Notes']
for ci, h in enumerate(headers3, 1):
    ws3.cell(row=3, column=ci, value=h)
style_header_row(ws3, 3, len(headers3))

# Fed: report states 100bp cumulative cut from 5.25-5.50% to 4.25-4.50%
# ECB: rates decreased to 2.75/2.90/3.15% effective 5 Feb 2025; previous not explicitly stated
# But we know from public record: before Feb 2025 cut, deposit was 3.00% (cut Dec 2024).
# The report does NOT explicitly state "from 3.00%" so we note this.

policy_data = [
    ['United States', 'Fed Funds Target Range (upper bound)',
     '4.50%', '5.50%', -100, 'Dec 2024 (last change)', 
     'FOMC lowered target range by cumulative 100bp at Sep, Nov, Dec 2024 meetings (Feb 2025 report). Maintained at 4.25-4.50% through Jun 2025 (June 2025 report).'],
    ['United States', 'Fed Funds Target Range (lower bound)',
     '4.25%', '5.25%', -100, 'Dec 2024 (last change)',
     'See above. Range unchanged since Dec 2024.'],
    ['Euro Area', 'ECB Deposit Facility Rate',
     '2.75%', '3.00% (inferred)', -25, '5 Feb 2025 (effective)',
     'ECB Bulletin states rates "were decreased to 2.75%... with effect from 5 February 2025." Previous level of 3.00% is inferred (not explicitly stated in the report text); the standard ECB cut increment of 25bp is assumed.'],
    ['Euro Area', 'ECB Main Refinancing Operations Rate',
     '2.90%', '3.15% (inferred)', -25, '5 Feb 2025 (effective)',
     'See above. Previous level inferred; same 25bp cut as deposit facility.'],
    ['Euro Area', 'ECB Marginal Lending Facility Rate',
     '3.15%', '3.40% (inferred)', -25, '5 Feb 2025 (effective)',
     'See above. Previous level inferred; same 25bp cut.'],
]

for ri, row in enumerate(policy_data, 4):
    for ci, val in enumerate(row, 1):
        ws3.cell(row=ri, column=ci, value=val)

style_data_rows(ws3, 4, 3 + len(policy_data), len(headers3))

col_widths_3 = [18, 36, 18, 20, 14, 24, 72]
for ci, w in enumerate(col_widths_3, 1):
    ws3.column_dimensions[get_column_letter(ci)].width = w

# ═══════════════════════════════════════════════════
# Sheet 4 — Data Notes
# ═══════════════════════════════════════════════════
ws4 = wb.create_sheet("Data Notes")

ws4.merge_cells('A1:B1')
ws4['A1'] = "Data Notes — Extraction Details, Assumptions & Clarifications"
ws4['A1'].font = title_font
ws4.row_dimensions[1].height = 28

headers4 = ['Category', 'Note']
for ci, h in enumerate(headers4, 1):
    ws4.cell(row=3, column=ci, value=h)
style_header_row(ws4, 3, len(headers4))

notes = [
    ['Source Reports',
     '1) Federal Reserve Monetary Policy Report (February 7, 2025) — data through ~Feb 4, 2025 (daily) / Dec 2024 (monthly) / 2024:Q4 (quarterly). '
     '2) Federal Reserve Monetary Policy Report (June 18, 2025) — data through Jun 16, 2025 (daily) / May 2025 (monthly) / 2025:Q1 (quarterly). '
     '3) ECB Economic Bulletin, Issue 1/2025 — cut-off date Jan 29, 2025 (GDP: Jan 30, 2025).'],
    ['US Inflation Measure',
     'The Fed targets PCE inflation (2% objective). Headline PCE = total PCE price index. Core PCE = PCE excluding food and energy. Data from BEA.'],
    ['Euro Area Inflation Measure',
     'The ECB targets HICP inflation (2% symmetric target). Headline HICP reported. "Core HICP" (excl. food & energy) is not reported as a single number in the text; component breakdowns are given: services 4.0%, food 2.6%, non-energy industrial goods 0.5% (Dec 2024).'],
    ['Euro Area GDP',
     'The advance estimate for Q4 2024 GDP was published on Jan 30, 2025 — one day after the ECB Bulletin cut-off date. Therefore no explicit GDP number appears in the report text.'],
    ['Euro Area Unemployment Rate',
     'The ECB Bulletin Chart 5 references the unemployment rate through November 2024, but the exact value is not quoted in the extracted text.'],
    ['Fed Policy Rate Changes',
     'The Feb 2025 Fed report explicitly states: "the FOMC lowered the target range for the policy rate by a cumulative 100 basis points over its September, November, and December meetings, bringing it to the current range of 4¼ to 4½ percent." '
     'The June 2025 report confirms the range has been maintained since the beginning of the year. Previous range computed as 5.25-5.50% (= 4.25-4.50% + 100bp).'],
    ['ECB Policy Rate Changes',
     'The ECB Bulletin (Issue 1/2025) states rates were "decreased to 2.75%, 2.90% and 3.15% respectively, with effect from 5 February 2025" but does NOT explicitly state the previous levels or the size of the cut. '
     'Previous levels of 3.00%, 3.15%, and 3.40% are inferred based on the ECB\'s standard 25bp increment and the prior policy trajectory. This assumption is clearly flagged in the Policy Rates sheet.'],
    ['US Core PCE (Dec 2024)',
     'The Feb 2025 Fed report does not give an explicit core PCE YoY number for Dec 2024 in the extracted text. However, the June 2025 report references "the 2.9 percent increase observed at the end of last year" for core PCE, allowing us to backfill the Dec 2024 core PCE as 2.9%.'],
    ['PCE Energy Prices',
     'The June 2025 Fed report states "PCE energy prices declined, on net, during the early part of this year, with the 12-month change through April indicating a drop of almost 6 percent." Recorded as approximately -6.0%.'],
    ['Inflation Target Deviation',
     'Deviation = latest headline inflation minus 2.0% target. For the US: PCE 2.1% - 2.0% = +0.1 pp (April 2025). For the euro area: HICP 2.4% - 2.0% = +0.4 pp (December 2024) / HICP 2.5% - 2.0% = +0.5 pp (January 2025 flash).'],
    ['General Caveat',
     'All values are taken exclusively from the extracted text of the three reports. Where values are inferred or approximated, this is explicitly noted. No external data sources were consulted.'],
]

for ri, row in enumerate(notes, 4):
    for ci, val in enumerate(row, 1):
        ws4.cell(row=ri, column=ci, value=val)

style_data_rows(ws4, 4, 3 + len(notes), len(headers4))

ws4.column_dimensions['A'].width = 28
ws4.column_dimensions['B'].width = 110

# Format all data rows in notes for readability
for r in range(4, 4 + len(notes)):
    ws4.row_dimensions[r].height = 72

# Save
wb.save('Macro_Snapshot_2025.xlsx')
print("Workbook 'Macro_Snapshot_2025.xlsx' created successfully.")
print("\nSheets:", wb.sheetnames)
print("\nKey Indicators rows:", len(key_data))
print("Inflation Gap rows:", len(gap_data))
print("Policy Rates rows:", len(policy_data))
print("Data Notes rows:", len(notes))

# Print extracted values for verification
print("\n── Extracted Values Verification ──")
for row in key_data:
    if row[2] is not None:
        print(f"  {row[0]:15s} | {row[1]:55s} | {str(row[2]):>8s} | {row[3]:20s} | {row[4]}")
    else:
        print(f"  {row[0]:15s} | {row[1]:55s} | {'N/A':>8s} | {row[3]:20s} | {row[4]}")

print("\n── Inflation Gaps ──")
for row in gap_data:
    print(f"  {row[0]:25s} | {row[1]:5.1f}% | Target: {row[2]:.1f}% | Deviation: {row[3]:+.1f} pp")

print("\n── Policy Rate Changes ──")
for row in policy_data:
    print(f"  {row[0]:15s} | {row[1]:42s} | {row[2]:>8s} | Prev: {row[3]:>8s} | Change: {row[4]:+d} bp")
