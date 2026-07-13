
import openpyxl

wb = openpyxl.load_workbook('Macro_Snapshot_2025.xlsx')
print(f"Sheets: {wb.sheetnames}")
print()

for name in wb.sheetnames:
    ws = wb[name]
    print(f"--- {name} ---")
    print(f"  Rows: {ws.max_row}, Cols: {ws.max_column}")
    # Print first 3 rows
    for row in ws.iter_rows(min_row=1, max_row=min(3, ws.max_row), values_only=True):
        print(f"  {[str(c)[:60] if c else '' for c in row]}")
    print()

# Verify key numbers
ws1 = wb['Key Indicators']
print("=== SPOT CHECKS ===")
# Check US headline PCE
for row in ws1.iter_rows(min_row=2, values_only=True):
    if row[0] and 'United States' in str(row[0]) and 'Headline' in str(row[1]):
        print(f"US Headline PCE: {row[2]} ({row[3]}), as of {row[4]}")
        break

ws2 = wb['Inflation Gap']
for row in ws2.iter_rows(min_row=2, values_only=True):
    print(f"Inflation Gap: {row[0]} | {row[1]}% vs {row[2]}% target = {row[3]} pp")

ws3 = wb['Policy Rates']
for row in ws3.iter_rows(min_row=2, values_only=True):
    print(f"Policy: {row[0]} | {row[1]} | Current: {row[2]} | Change: {row[4]}")

print("\nVerification complete.")
